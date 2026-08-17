from __future__ import annotations

import math
import tempfile
from pathlib import Path
from typing import Any, Mapping, Sequence

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.io import loadmat

try:
    from .synchronized_mat_export import export_matlab_simulation
except ImportError:
    from synchronized_mat_export import export_matlab_simulation


EXCEL_MAX_DATA_ROWS = 1_048_575
EXCEL_NO_LIMIT_SENTINEL = 65_535.0

# Positive infinity has a useful physical meaning for these channels: no finite
# curve restriction / effectively straight road. Excel should still receive a
# numeric value in every row, so use one documented finite sentinel instead of
# silently writing an empty cell.
_INFINITY_SENTINEL_SIGNALS = {
    "curve_radius_m",
    "v_curve_limit_kmh",
}

_PREFERRED_INPUT_ORDER = (
    "sample_index",
    "time_s",
    "distance_m",
    "lat_deg",
    "lon_deg",
    "elevation_m",
    "grade_pct",
    "curve_radius_m",
    "v_kmh",
    "v_target_kmh",
    "a_mps2",
    "v_road_limit_kmh",
    "v_surface_limit_kmh",
    "v_curve_limit_kmh",
    "v_base_target_kmh",
    "v_planned_kmh",
    "v_actual_kmh",
    "v_noise_kmh",
    "speed_policy_limit_kmh",
    "speeding_over_kmh",
    "speeding_points",
    "p_total_kw",
    "p_acceleration_kw",
    "p_grade_kw",
    "p_rolling_kw",
    "p_air_kw",
    "p_trailer_kw",
    "p_traction_kw",
    "p_recuperation_kw",
    "e_traction_cum_kwh",
    "e_recuperation_cum_kwh",
    "e_net_cum_kwh",
)


def _excel_value(value: Any) -> Any:
    if isinstance(value, np.generic):
        value = value.item()
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    return number if math.isfinite(number) else None


def _flat_numeric(value: Any) -> np.ndarray:
    try:
        return np.asarray(value, dtype=np.float64).reshape(-1)
    except (TypeError, ValueError):
        return np.empty(0, dtype=np.float64)


def _ordered_input_names(inputs: Mapping[str, np.ndarray]) -> list[str]:
    preferred = [name for name in _PREFERRED_INPUT_ORDER if name in inputs]
    preferred_set = set(preferred)
    remaining = sorted(name for name in inputs if name not in preferred_set)
    return preferred + remaining


def _normalize_simulation_inputs(
    inputs: Mapping[str, np.ndarray],
) -> tuple[dict[str, np.ndarray], dict[str, int]]:
    """Return one complete numeric Excel table on the input_time_s master grid.

    Length mismatches and unexpected NaN/inf values are export errors. This is
    intentional: a visibly failed export is much safer than a workbook where one
    signal silently ends earlier than the others. Only +inf in the documented
    no-curve-limit channels is converted to a finite numeric sentinel.
    """

    if "time_s" not in inputs:
        raise ValueError("Der synchronisierte Excel-Export enthält keinen Zeitvektor input_time_s.")

    n = int(np.asarray(inputs["time_s"]).size)
    if n <= 0:
        raise ValueError("Der synchronisierte Excel-Export enthält keine Zeitschritte.")

    invalid_lengths = {
        name: int(np.asarray(values).size)
        for name, values in inputs.items()
        if np.asarray(values).size != n
    }
    if invalid_lengths:
        raise ValueError(
            "Excel-Simulationssignale haben unterschiedliche Längen. "
            f"Masterlänge input_time_s={n}; abweichend: {invalid_lengths}"
        )

    normalized: dict[str, np.ndarray] = {}
    substitutions: dict[str, int] = {}
    for name, values in inputs.items():
        array = np.asarray(values, dtype=np.float64).reshape(-1).copy()
        nan_count = int(np.count_nonzero(np.isnan(array)))
        neg_inf_count = int(np.count_nonzero(np.isneginf(array)))
        pos_inf = np.isposinf(array)
        pos_inf_count = int(np.count_nonzero(pos_inf))

        if nan_count or neg_inf_count:
            raise ValueError(
                f"Excel-Signal '{name}' enthält {nan_count} NaN und "
                f"{neg_inf_count} -inf. Der Export wird abgebrochen, damit keine "
                "unbemerkten leeren Zellen entstehen."
            )

        if pos_inf_count:
            if name not in _INFINITY_SENTINEL_SIGNALS:
                raise ValueError(
                    f"Excel-Signal '{name}' enthält {pos_inf_count} +inf-Werte. "
                    "Für diesen Kanal ist kein Excel-Ersatzwert definiert."
                )
            array[pos_inf] = EXCEL_NO_LIMIT_SENTINEL
            substitutions[name] = pos_inf_count
        else:
            substitutions[name] = 0

        if not np.all(np.isfinite(array)):
            raise ValueError(f"Excel-Signal '{name}' ist nach der Normalisierung nicht vollständig numerisch.")
        normalized[name] = array

    return normalized, substitutions


def _header_style(sheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _fit_columns(sheet, *, max_width: int = 34) -> None:
    for index, column in enumerate(sheet.iter_cols(), start=1):
        width = 10
        for cell in column[:200]:
            if cell.value is not None:
                width = max(width, len(str(cell.value)) + 2)
        sheet.column_dimensions[get_column_letter(index)].width = min(max_width, width)


def _write_key_value_sheet(workbook: Workbook, title: str, values: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(["Parameter", "Wert"])
    for key in sorted(values):
        value = values[key]
        if isinstance(value, (Mapping, list, tuple, np.ndarray)):
            continue
        sheet.append([str(key), _excel_value(value)])
    _header_style(sheet)
    _fit_columns(sheet)


def _write_information_sheet(
    workbook: Workbook,
    names: Sequence[str],
    inputs: Mapping[str, np.ndarray],
    substitutions: Mapping[str, int],
) -> None:
    sheet = workbook.create_sheet("Information")
    sheet.append(
        [
            "Spalte",
            "Datenkanal",
            "Anzahl Werte",
            "Leere Zellen",
            "Ersetzte +inf",
            "Datentyp",
        ]
    )
    for index, name in enumerate(names, start=1):
        values = np.asarray(inputs[name], dtype=np.float64).reshape(-1)
        sheet.append(
            [
                index,
                name,
                int(values.size),
                0,
                int(substitutions.get(name, 0)),
                "double",
            ]
        )
    sheet.append([])
    sheet.append(
        [
            None,
            "Hinweis",
            None,
            None,
            None,
            f"+inf in Kurvenkanälen wird als {EXCEL_NO_LIMIT_SENTINEL:g} exportiert.",
        ]
    )
    _header_style(sheet)
    _fit_columns(sheet, max_width=48)


def _record_rows(records: Sequence[Any]) -> tuple[list[str], list[list[Any]]]:
    rows = [row for row in records if isinstance(row, Mapping)]
    if not rows:
        return [], []
    keys = sorted(
        {
            str(key)
            for row in rows
            for key, value in row.items()
            if not isinstance(value, (Mapping, list, tuple, np.ndarray))
        }
    )
    return keys, [[_excel_value(row.get(key)) for key in keys] for row in rows]


def _write_records_sheet(workbook: Workbook, title: str, records: Sequence[Any]) -> None:
    keys, rows = _record_rows(records)
    if not keys:
        return
    sheet = workbook.create_sheet(title)
    sheet.append(keys)
    for row in rows:
        sheet.append(row)
    _header_style(sheet)
    _fit_columns(sheet)


def _write_drivers_sheet(
    workbook: Workbook,
    parameters: Mapping[str, Any] | None,
    comparison: Mapping[str, Any] | None,
) -> None:
    records: list[dict[str, Any]] = []
    if isinstance(parameters, Mapping):
        records.append({"name": "Aktuell", **dict(parameters)})
    if isinstance(comparison, Mapping):
        configs = comparison.get("configs", [])
        if isinstance(configs, (list, tuple)):
            for config in configs:
                if not isinstance(config, Mapping):
                    continue
                row: dict[str, Any] = {"name": str(config.get("name", "Vergleich"))}
                config_parameters = config.get("parameters", {})
                if isinstance(config_parameters, Mapping):
                    row.update(config_parameters)
                records.append(row)
    if not records:
        return

    all_keys = sorted(
        {
            str(key)
            for row in records
            for key, value in row.items()
            if not isinstance(value, (Mapping, list, tuple, np.ndarray))
        }
        - {"name"}
    )
    sheet = workbook.create_sheet("Drivers")
    sheet.append(["name", *all_keys])
    for row in records:
        sheet.append(
            [_excel_value(row.get("name")), *[_excel_value(row.get(key)) for key in all_keys]]
        )
    _header_style(sheet)
    _fit_columns(sheet)


def export_excel_simulation(
    result: Mapping[str, Any],
    output_path: str | Path,
    *,
    route: Mapping[str, Any] | None = None,
    parameters: Mapping[str, Any] | None = None,
    power_data: Mapping[str, Any] | None = None,
    elevation_m: np.ndarray | None = None,
    source_route: str | Path | None = None,
    source_dem: str | Path | None = None,
    comparison: Mapping[str, Any] | None = None,
) -> Path:
    """Export the synchronized MAT input contract as a complete numeric XLSX workbook."""

    path = Path(output_path).expanduser().resolve()
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)

    # Reuse the synchronized MAT contract so MAT and Excel can never implement
    # independent time-grid/resampling rules.
    with tempfile.TemporaryDirectory(prefix="gps_excel_export_") as temporary_directory:
        mat_path = export_matlab_simulation(
            result,
            Path(temporary_directory) / "synchronized.mat",
            route=route,
            parameters=parameters,
            power_data=power_data,
            elevation_m=elevation_m,
            source_route=source_route,
            source_dem=source_dem,
            comparison=comparison,
        )
        workspace = loadmat(mat_path, simplify_cells=True)

    raw_inputs: dict[str, np.ndarray] = {}
    for name, value in workspace.items():
        if not name.startswith("input_"):
            continue
        signal = _flat_numeric(value)
        if signal.size:
            raw_inputs[name[len("input_") :]] = signal
    if not raw_inputs or "time_s" not in raw_inputs:
        raise ValueError("Der synchronisierte Excel-Export enthält keinen Zeitvektor input_time_s.")

    inputs, substitutions = _normalize_simulation_inputs(raw_inputs)
    n = int(inputs["time_s"].size)
    if n > EXCEL_MAX_DATA_ROWS:
        raise ValueError(
            f"Die Simulation enthält {n:,} Zeitschritte. Excel unterstützt pro Tabellenblatt "
            f"höchstens {EXCEL_MAX_DATA_ROWS:,} Datenzeilen."
        )

    workbook = Workbook()
    simulation = workbook.active
    simulation.title = "Simulation"
    names = _ordered_input_names(inputs)
    simulation.append(names)
    for row_index in range(n):
        # The inputs have already been validated as finite numeric N-vectors, so
        # no simulation cell can silently become blank at this point.
        simulation.append([float(inputs[name][row_index]) for name in names])
    _header_style(simulation)
    _fit_columns(simulation, max_width=25)

    _write_information_sheet(workbook, names, inputs, substitutions)

    summary = result.get("summary", {})
    if isinstance(summary, Mapping):
        _write_key_value_sheet(workbook, "Summary", summary)
    if isinstance(parameters, Mapping):
        _write_key_value_sheet(workbook, "Parameters", parameters)
    _write_drivers_sheet(workbook, parameters, comparison)

    events = result.get("events", {})
    if isinstance(events, Mapping):
        traffic = events.get("traffic_lights", [])
        if isinstance(traffic, (list, tuple)):
            _write_records_sheet(workbook, "Traffic_Lights", traffic)

    if isinstance(route, Mapping):
        segments = route.get("segments", [])
        if isinstance(segments, (list, tuple)):
            _write_records_sheet(workbook, "Segments", segments)

    workbook.save(path)
    return path


__all__ = [
    "EXCEL_MAX_DATA_ROWS",
    "EXCEL_NO_LIMIT_SENTINEL",
    "export_excel_simulation",
]
