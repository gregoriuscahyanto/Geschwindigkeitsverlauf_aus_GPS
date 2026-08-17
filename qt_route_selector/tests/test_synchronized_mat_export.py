from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.io import loadmat

from qt_route_selector.synchronized_excel_export import (
    EXCEL_NO_LIMIT_SENTINEL,
    export_excel_simulation,
)
from qt_route_selector.synchronized_mat_export import (
    MAT_NO_LIMIT_SENTINEL,
    export_matlab_simulation,
)


class SynchronizedMatExportTests(unittest.TestCase):
    @staticmethod
    def _fixture() -> tuple[dict, np.ndarray, dict]:
        result = {
            "distance": {
                "distance_m": np.asarray([0.0, 30.0, 70.0, 100.0]),
                "latitude": np.asarray([48.0, 48.001, 48.002, 48.003]),
                "longitude": np.asarray([9.0, 9.001, 9.002, 9.003]),
                "curve_radius_m": np.asarray([np.inf, 80.0, 35.0, np.inf]),
                "road_limit_kmh": np.asarray([50.0, 50.0, 30.0, 70.0]),
                "surface_limit_kmh": np.asarray([50.0, 50.0, 30.0, 70.0]),
                "curve_limit_kmh": np.asarray([np.inf, 55.0, 32.0, np.inf]),
                "base_target_kmh": np.asarray([50.0, 50.0, 30.0, 70.0]),
                "planned_speed_kmh": np.asarray([0.0, 35.0, 28.0, 0.0]),
                "actual_speed_kmh": np.asarray([0.0, 34.0, 27.0, 0.0]),
                "noise_kmh": np.asarray([0.0, 0.2, -0.1, 0.0]),
                "post_curve_boost_kmh": np.asarray([0.0, 1.0, 2.0, 0.0]),
            },
            "time": {
                "time_s": np.asarray([0.0, 2.0, 4.0, 6.0, 8.0]),
                "distance_m": np.asarray([0.0, 18.0, 44.0, 73.0, 100.0]),
                "speed_kmh": np.asarray([0.0, 25.0, 31.0, 27.0, 0.0]),
                "target_kmh": np.asarray([0.0, 30.0, 35.0, 30.0, 0.0]),
                "acceleration_mps2": np.asarray([0.0, 1.2, 0.3, -0.8, -1.5]),
            },
            "events": {
                "traffic_lights": [{"distance_m": 44.0, "dwell_s": 2.0}],
                "traffic_light_dwell_intervals_s": [[4.0, 6.0]],
                "overtaking": [{"follow_start_m": 18.0, "pass_end_m": 73.0}],
            },
            "summary": {"distance_km": 0.1, "duration_min": 0.1333},
        }
        elevation = np.asarray([400.0, 405.0, 410.0, 408.0])
        power = {
            "grade_spatial": np.asarray([0.0, 0.04, -0.02, 0.0]),
            "total_kw": np.asarray([0.0, 12.0, 18.0, 7.0, -3.0]),
            "acceleration_kw": np.zeros(5),
            "grade_kw": np.zeros(5),
            "rolling_kw": np.zeros(5),
            "air_kw": np.zeros(5),
            "trailer_kw": np.zeros(5),
            "traction_power_kw": np.asarray([0.0, 12.0, 18.0, 7.0, 0.0]),
            "recuperation_power_kw": np.asarray([0.0, 0.0, 0.0, 0.0, 3.0]),
            "cumulative_traction_energy_kwh": np.zeros(5),
            "cumulative_recuperation_energy_kwh": np.zeros(5),
            "cumulative_net_energy_kwh": np.zeros(5),
            "traction_energy_kwh": 0.25,
        }
        return result, elevation, power

    def test_every_mat_variable_has_same_length_and_only_finite_values(self) -> None:
        result, elevation, power = self._fixture()
        parameters = {
            "driver_hard_max_kmh": 140.0,
            "temperament": 1.0,
            "simulation_seed": 42,
        }

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = export_matlab_simulation(
                result,
                Path(temporary_directory) / "sync.mat",
                parameters=parameters,
                power_data=power,
                elevation_m=elevation,
            )
            loaded = loadmat(path)

        variables = {
            name: np.asarray(value, dtype=float)
            for name, value in loaded.items()
            if not name.startswith("__")
        }
        self.assertTrue(variables)

        # New hard MAT contract: there are no structs, scalar metadata, route
        # arrays or event lists with another natural length. Every stored
        # variable is exactly one finite N x 1 signal.
        self.assertEqual({array.shape for array in variables.values()}, {(5, 1)})
        self.assertTrue(all(np.all(np.isfinite(array)) for array in variables.values()))

        required = {
            "time_s",
            "distance_m",
            "v_kmh",
            "curve_radius_m",
            "elevation_m",
            "grade_pct",
            "lat_deg",
            "lon_deg",
            "v_road_limit_kmh",
            "p_total_kw",
            "post_curve_boost_kmh",
            "traffic_light_stop",
            "traffic_light_active",
            "traffic_light_dwell_s",
            "overtaking_active",
            "param_driver_hard_max_kmh",
            "summary_distance_km",
            "input_time_s",
            "input_elevation_m",
        }
        self.assertTrue(required.issubset(variables.keys()), required - variables.keys())

        np.testing.assert_allclose(variables["time_s"].ravel(), [0, 2, 4, 6, 8])
        np.testing.assert_allclose(variables["v_kmh"].ravel(), [0, 25, 31, 27, 0])
        np.testing.assert_allclose(variables["elevation_m"].ravel(), [400.0, 403.0, 407.0, 409.8, 408.0])
        np.testing.assert_allclose(
            variables["param_driver_hard_max_kmh"].ravel(),
            np.full(5, 140.0),
        )
        self.assertEqual(variables["curve_radius_m"][0, 0], MAT_NO_LIMIT_SENTINEL)
        self.assertEqual(variables["curve_radius_m"][-1, 0], MAT_NO_LIMIT_SENTINEL)
        self.assertEqual(variables["v_curve_limit_kmh"][0, 0], MAT_NO_LIMIT_SENTINEL)
        self.assertEqual(variables["v_curve_limit_kmh"][-1, 0], MAT_NO_LIMIT_SENTINEL)

    def test_mat_refuses_completely_missing_required_elevation(self) -> None:
        result, _elevation, power = self._fixture()
        missing_elevation = np.full(4, np.nan)

        with tempfile.TemporaryDirectory() as temporary_directory:
            with self.assertRaisesRegex(ValueError, "elevation_m"):
                export_matlab_simulation(
                    result,
                    Path(temporary_directory) / "broken.mat",
                    power_data=power,
                    elevation_m=missing_elevation,
                )

    def test_excel_simulation_sheet_uses_same_five_sample_grid_without_blanks(self) -> None:
        result, elevation, power = self._fixture()
        route = {
            "segments": [
                {"from_index": 0, "to_index": 1, "distance_m": 100.0, "maxspeed_kmh": 50.0}
            ]
        }
        parameters = {"driver_profile": "normalo", "driver_hard_max_kmh": 140.0}

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = export_excel_simulation(
                result,
                Path(temporary_directory) / "sync.xlsx",
                route=route,
                parameters=parameters,
                power_data=power,
                elevation_m=elevation,
            )
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook["Simulation"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            rows = list(sheet.iter_rows(min_row=2, values_only=True))

            self.assertEqual(len(rows), 5)
            required = {
                "time_s",
                "distance_m",
                "v_kmh",
                "curve_radius_m",
                "elevation_m",
                "grade_pct",
                "p_total_kw",
            }
            self.assertTrue(required.issubset(set(headers)), required - set(headers))
            self.assertIn("Information", workbook.sheetnames)
            self.assertIn("Summary", workbook.sheetnames)
            self.assertIn("Parameters", workbook.sheetnames)
            self.assertIn("Drivers", workbook.sheetnames)
            self.assertIn("Traffic_Lights", workbook.sheetnames)
            self.assertIn("Segments", workbook.sheetnames)

            self.assertTrue(rows)
            for row in rows:
                self.assertEqual(len(row), len(headers))
                self.assertTrue(all(value is not None for value in row), row)

            time_column = headers.index("time_s")
            speed_column = headers.index("v_kmh")
            elevation_column = headers.index("elevation_m")
            radius_column = headers.index("curve_radius_m")
            curve_limit_column = headers.index("v_curve_limit_kmh")
            self.assertEqual([row[time_column] for row in rows], [0, 2, 4, 6, 8])
            self.assertEqual([row[speed_column] for row in rows], [0, 25, 31, 27, 0])
            self.assertIsNotNone(rows[-1][elevation_column])
            self.assertEqual(rows[-1][elevation_column], 408.0)
            self.assertEqual(rows[0][radius_column], EXCEL_NO_LIMIT_SENTINEL)
            self.assertEqual(rows[-1][radius_column], EXCEL_NO_LIMIT_SENTINEL)
            self.assertEqual(rows[0][curve_limit_column], EXCEL_NO_LIMIT_SENTINEL)
            self.assertEqual(rows[-1][curve_limit_column], EXCEL_NO_LIMIT_SENTINEL)

            info = workbook["Information"]
            info_rows = list(info.iter_rows(min_row=2, values_only=True))
            channel_rows = [row for row in info_rows if isinstance(row[0], int)]
            self.assertEqual(len(channel_rows), len(headers))
            self.assertTrue(all(row[2] == 5 for row in channel_rows))
            self.assertTrue(all(row[3] == 0 for row in channel_rows))
            workbook.close()

    def test_excel_refuses_stale_shorter_time_series_instead_of_padding_blanks(self) -> None:
        result, elevation, power = self._fixture()
        power = dict(power)
        power["total_kw"] = np.asarray([0.0, 12.0, 18.0, 7.0])

        with tempfile.TemporaryDirectory() as temporary_directory:
            with self.assertRaisesRegex(ValueError, "p_total_kw"):
                export_excel_simulation(
                    result,
                    Path(temporary_directory) / "broken.xlsx",
                    power_data=power,
                    elevation_m=elevation,
                )


if __name__ == "__main__":
    unittest.main()
